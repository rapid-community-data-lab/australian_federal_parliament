"""
This script creates an index of the first speeches from the Parliamentary Handbook.

This is just links to the specific first speeches, with additional contextual
information about the speaker for sorting and filtering. Note that not all
parliamentarians have recognised first speeches in the handbook.

python first_speeches_index.py

# /// script
# requires-python = ">=3.12"
# dependencies = [
#   "openpyxl"
# ]
# ///

"""

import datetime
import sqlite3

from openpyxl import Workbook, styles
from openpyxl.utils.cell import get_column_letter

# Note - isolation_level=None won't work in future versions of Python (sometime after
# 3.13), this should be using the autocommit=True value instead, but requires some
# investigation for different python versions
db = sqlite3.connect("../../oz_federal_hansard.db", isolation_level=None)

# Create a temporary table to parse the speech dates properly.
db.executescript("""
    CREATE temporary table first_speeches (
        phid,
        title,
        link,
        speech_date
    )
    """)

speeches = db.execute("SELECT * from parliamentarian_speeches")

for phid, title, link, _ in speeches:

    if not title.startswith("First"):
        continue

    date_component = title.split(" ")[-1]

    try:
        components = [int(comp) for comp in date_component.split(".") if comp]
        day, month, year = components
    except Exception as e:
        # Override for B36 with a mangled date with a space
        if len(components) == 1 and components[0] == 2008 and phid == "B36":
            day, month, year = (18, 2, 2008)
        else:
            print(phid, title, e)
            raise

    dt = datetime.date(year, month, day)

    db.execute(
        "insert into first_speeches values(?, ?, ?, ?)",
        (phid, title, link, dt.isoformat()),
    )


query = """
    
    SELECT 
        first_speeches.*,
        parliamentarian.display_name,
        parliamentarian.gender,
        parliamentarian.date_of_birth,
        parliamentarian.date_of_death,
        substr(timediff(speech_date, parliamentarian.date_of_birth), 4, 2)
            as age_at_speech,
        party.name as party_name
    from first_speeches
    left outer join parliamentarian using(phid)
    left outer join party_member pm on pm.phid = parliamentarian.phid
        and first_speeches.speech_date between pm.start_date and 
            coalesce(pm.end_date, '3000-01-01')
    left outer join party using(party_id)
    order by speech_date

"""

workbook = Workbook()
worksheet = workbook.active

results = db.execute(query)

header = [col[0] for col in results.description]

worksheet.append(header)

for row in results:
    row = list(row)
    worksheet.append(row)

# Update transcript link to be a proper hyperlink
all_rows = worksheet.rows
next(all_rows)  # skip header

for row in all_rows:
    link = row[2].value

    row[2].hyperlink = link
    row[2].value = "Speech Link"

# Zebra stripe speeches and set text to wrap
all_rows = worksheet.rows
next(all_rows)  # skip header

colour = True
last_speech = (None, None, None)

solid_fill = styles.PatternFill(fill_type="solid", fgColor="efefef")

for row in all_rows:

    current_speech = (row[0].value, row[1].value, row[4].value)

    if current_speech != last_speech:
        colour = not colour
        last_speech = current_speech

    if colour:
        for cell in row:
            cell.fill = solid_fill

    for cell in row:
        cell.alignment = styles.Alignment(
            wrap_text=True, vertical="top", horizontal="left"
        )

# Format column widths and alignments for readability
for i, header in enumerate(header):
    col = worksheet.column_dimensions[get_column_letter(i + 1)]

    col.width = 15

    if header == "paragraph_text":
        col.width = 40

# Freeze the header

worksheet.freeze_panes = "A2"

workbook.save("first_speeches_index.xlsx")
