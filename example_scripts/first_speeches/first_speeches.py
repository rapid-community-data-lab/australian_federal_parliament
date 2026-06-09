"""
This script is an exploration of the 'first.speech' metadata attached to some
talk.start/talker tags in the XML transcripts. This turns out to be more of a negative
result in that it actually identifies that the first.speech flag is not consistently
applied except in a small window of time for the XML transcripts.

python best_country.py

# /// script
# requires-python = ">=3.12"
# dependencies = [
#   "openpyxl"
# ]
# ///

"""

import datetime
import sqlite3

from openpyxl import Workbook, styles
from openpyxl.utils.cell import get_column_letter

# Note - isolation_level=None won't work in future versions of Python (sometime after
# 3.13), this should be using the autocommit=True value instead, but requires some
# investigation for different python versions
db = sqlite3.connect("../../oz_federal_hansard.db", isolation_level=None)


# Create a temporary table to handle interjections as CSS classes - this is a temporary
# measure until we've resolved all of the speaker identification issues.
db.executescript("""
    CREATE temporary table contains_interjection (
        para_id integer primary key,
        interjection bool
    );

    insert into contains_interjection
        select distinct
            para_id, 
            1
        from paragraph_enclosed_class
        where lower(class) like '%inter%';
    """)

query = """
    WITH matching_speech as (
        /* All fragments with the first_speech set to '1'*/
        select distinct
            session_id, 
            fragment_number 
        from fragment_speaker_info
        where first_speech = '1'
    )
    select 
        session.date,
        session.chamber,
        session.transcript_pdf_url as full_transcript_link,
        debate.title as debate_title,
        fragment_number as speech_number,
        case when interjection then 'interjector' else parliamentarian.display_name end
            as speaker,
        case when interjection then null else party.name end as party,
        case when interjection then null else parliamentarian.gender end
            as gender,
        case when interjection then null else parliamentarian.date_of_birth end
            as date_of_birth,
        case when interjection then null else 
            -- time diff as an iso8601 like string, then just pulling the years out.
            substr(timediff(session.date, parliamentarian.date_of_birth), 4, 2) end
            as age,
        paragraph.paragraph_text || '\n' as paragraph_text
    from paragraph
    inner join session using(session_id)
    inner join debate using(debate_id)
    left outer join parliamentarian on speaker_id = parliamentarian.phid
    left outer join party_member on speaker_id = party_member.phid
        and session.date between party_member.start_date and 
            coalesce(party_member.end_date, '3000-01-01')
    left outer join party using(party_id)
    left outer join contains_interjection using(para_id)
    -- Find interjections using the classes assigned to paragraphs, as a workaround
    -- in advance of figuring out all the nuances of speakers...
    where (paragraph.session_id, fragment_number) in (
        select session_id, fragment_number from matching_speech
    )
    order by session.date, chamber, speech_number

"""

workbook = Workbook()
worksheet = workbook.active

results = db.execute(query)

header = [col[0] for col in results.description]

worksheet.append(header)

for row in results:
    row = list(row)
    # Save date as a proper datetime.
    date = row[0]
    row[0] = datetime.date.fromisoformat(date)

    worksheet.append(row)

# Update transcript link to be a proper hyperlink
all_rows = worksheet.rows
next(all_rows)  # skip header

for row in all_rows:
    link = row[2].value

    row[2].hyperlink = link
    row[2].value = "Session Transcript"

# Zebra stripe speeches and set text to wrap
all_rows = worksheet.rows
next(all_rows)  # skip header

colour = True
last_speech = (None, None, None)

solid_fill = styles.PatternFill(fill_type="solid", fgColor="efefef")

for row in all_rows:

    current_speech = (row[0].value, row[1].value, row[4].value)

    if current_speech != last_speech:
        colour = not colour
        last_speech = current_speech

    if colour:
        for cell in row:
            cell.fill = solid_fill

    for cell in row:
        cell.alignment = styles.Alignment(
            wrap_text=True, vertical="top", horizontal="left"
        )

# Format column widths and alignments for readability
for i, header in enumerate(header):
    col = worksheet.column_dimensions[get_column_letter(i + 1)]

    col.width = 15

    if header == "paragraph_text":
        col.width = 40

# Freeze the header

worksheet.freeze_panes = "A2"

workbook.save("first_speeches.xlsx")
